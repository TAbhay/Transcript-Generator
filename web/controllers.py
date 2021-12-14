import os
import csv
from typing import ParamSpec
import openpyxl
from openpyxl.cell.read_only import EmptyCell
from openpyxl.styles import PatternFill  # Connect cell styles
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, Fill  # Connect styles for text
from openpyxl.styles import colors  # Connect colors for text and cells
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import smtplib
from email.message import EmailMessage
import glob
import re
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4 ,landscape
from reportlab.lib.pagesizes import A3 ,landscape
from reportlab.platypus import Table
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus  import Table, Image
from web import btechMarksheet
from web import mtechMarksheet
#from reportlab.pdfbase.ttfonts import TTfont
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.pdfmetrics import registerFontFamily
from reportlab.pdfbase.ttfonts import TTFont

# Registered font family
pdfmetrics.registerFont(TTFont('Vera', 'Vera.ttf'))
pdfmetrics.registerFont(TTFont('VeraBd', 'VeraBd.ttf'))
pdfmetrics.registerFont(TTFont('VeraIt', 'VeraIt.ttf'))
pdfmetrics.registerFont(TTFont('VeraBI', 'VeraBI.ttf'))
# Registered fontfamily
registerFontFamily('Vera',normal='Vera',bold='VeraBd',italic='VeraIt',boldItalic='VeraBI')  
import os
import csv
import openpyxl
import re
##  Helper Functions
## Please refer below to understand the approach of the problem
## Declaring constanst for the uses

GRADE = {"AA":10,"AB":9,"BB":8,"BC":7,"CC":6,"CD":5,"DD":4,"F":0,"I":0,} #Grades
Header = ["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"] # Header for sem results
OverAllHeader = ["Roll No.","Name of Student","Discipline","Semester No.","Semester wise Credit Taken","SPI","Total Credits Taken","CPI"]   # Header for overall result
overallresult = ""
def mapSubject():    # helper function to map subject and subject details
    try:
        subject = {}
        j = 0
        with open("media/input/subjects_master.csv",'r',newline='') as File:
            reader = list(csv.reader(File))
            for i in range(1,len(reader)):
                row = reader[i]
                key = row[0]
                temp = [row[1],row[2],row[3]]
                subject[key] = temp 
        return subject
    except FileNotFoundError:
        print("Oops ! subjects_master.csv file not found or Please enter valid path")         # if file does not exist or not valid path
        return
    
def mapNameAndDiscipline():  # helper function to map rollNo with name & discipline
    try: 
        name = {}
        j = 0 
        with open("media/input/names-roll.csv",'r',newline='') as File:
            reader = list(csv.reader(File))
            for i in range(1,len(reader)):
                row = reader[i]
                rollNo = row[0]
                discipline = "".join(re.split("[^a-zA-Z]*", rollNo))
                temp = [row[1],discipline]
                name[rollNo] = temp 
        return name
    except FileNotFoundError:
        print("Oops ! names-roll.csv file not found or Please enter valid path")         # if file does not exist or not valid path
        return
def getSemDataByRollNo(subject,name):   # helper function to extract data from grades.csv and map it with rollNo and name
    
    try:
        result = {}      # dictionary to store subject as key
        j =0
        with open("media/input/grades.csv",'r',newline='') as File:         #the csv file is stored in a File object
                reader= list(csv.reader(File))                             #csv.reader is used to read a file
                for i in range(1,len(reader)):
                    row = reader[i]
                    rollNo = row[0]      # using rollno as key
                    sem = row[1]
                    sub = subject[row[2]]
                    temp = [row[2],sub[0],sub[1],sub[2],row[5],row[4]]  
                    if rollNo in result:                            
                        if sem in result[rollNo]:
                            result[rollNo][sem].append(temp)
                        else:
                            result[rollNo][sem] = []
                            result[rollNo][sem].append(temp)
                    else:
                        result[rollNo] = {}
                        result[rollNo]["Overall"] = {"1":rollNo,"2":name[rollNo][0],"3":name[rollNo][1]}
                        result[rollNo][sem] = []
                        result[rollNo][sem].append(temp)
        return result      
    except FileNotFoundError:
        print("Oops ! grades.csv file not found or Please enter valid path")         # if file does not exist or not valid path
        return

def generateResult(rollNo,semResult):   # helper function to calculate the cpi , spi etc
    try:
        allSemRes = semResult[rollNo]
        spi = []
        cred = []
        cpi  =  []
        semNum = []
        credSum =[]
        totalCredGrades = 0
        totalCredits = 0
        for sem in range(1,11):
            sem  = str(sem)
            if not sem in allSemRes:
                continue
            semRes = allSemRes[sem]
            semNum.append(sem)
            credit  = 0
            credGrades = 0
            grades = 0
            for sub in semRes:
                credit = credit + int(sub[3])
                credGrades = credGrades + int(sub[3])*GRADE["".join(re.split("[^a-zA-Z]*", sub[5]))]
            totalCredGrades = totalCredGrades + credGrades
            totalCredits = totalCredits + credit
            cred.append(credit)
            credSum.append(totalCredits)
            spi.append(round( credGrades/credit,2))
            cpi.append(round(totalCredGrades/totalCredits,2))
            #semRes.append([credit,credGrades,round( credGrades/credit,2)],round(totalCredGrades/totalCredits,2))
        overall  = allSemRes['Overall']
        overall['4'] = semNum
        overall['5'] = cred
        overall['6'] = spi
        overall['7'] = credSum
        overall['8'] = cpi
    except:
        print("Something occured while generating result")

def generateOutput(rollNo,result):   # writing to sheet
    try:
        filePath = "output/"+rollNo+".xlsx" 
        wb = openpyxl.Workbook() 
        sheet = wb.active
        sheet.title = "Overall" 
        rw = 1
        for item in OverAllHeader:       # generating overall result
            sheet.cell(row = rw,column = 1).value = item 
            rw = rw + 1
        overAll = result['Overall']
        sheet.cell(row = 1,column = 2).value = overAll['1']
        sheet.cell(row = 2,column = 2).value = overAll['2']
        sheet.cell(row = 3,column = 2).value = overAll['3']
        for i in range(4,9):
            value  = overAll[str(i)]
            col = 2
            for item in value:
                sheet.cell(row = i,column = col).value = item      
                col = col + 1
        for i in range(1,11):
            sem = str(i)
            if not sem in result:
                continue
            col = 1
            sheet = wb.create_sheet("Sem"+sem)     # generating result for each sem
            for item in Header:
                sheet.cell(row = 1,column = col).value = item       
                col = col + 1
            r = 2
            
            for row in result[sem]:
                sheet.cell(row=r, column=1).value = r-1
                col = 2
                for item in row:                                    
                    sheet.cell(row=r, column=col).value = item
                    col = col + 1
                r = r + 1
        wb.save(filePath)
    except:
        print(rollNo)
        print("Error in writing in workbook . Please have a look")
        return
          
"""
Main function 
"""    
def generate_marksheet():
    try:
        print("Program running !")
        if not os.path.isdir("output"):   # if folder does not exist , creating one
           os.mkdir("output")  
        subject = mapSubject()               # subject map
        name    = mapNameAndDiscipline()     # name and details map
        semResult = getSemDataByRollNo(subject,name)  # semester data extracted
        #print(semResult)

        for rollNo,sem in semResult.items():   # processig semResult data and generating cpi ,spi etc
             generateResult(rollNo,semResult)
        overallresult = rollNo
        # for rollNo,sem in semResult.items():  # writing the results in the sheet
        #     generateOutput(rollNo,semResult[rollNo])
        return semResult
        print("Program ran successfully ! Check output folder")      
    except:
        print("Oops ! Something occurred")         # if file does not exist or not valid path
        return

def overAllmarksheet():
    try:
        overallresult = generate_marksheet()
        for key ,value in overallresult.items():
            typeofProgram = str(key[2]+key[3])
            if typeofProgram=="01":
                if len(value)>9:
                    continue
                btechMarksheet.semesterTables({key:value})
            else:
                if len(value)>7:
                    continue
                mtechMarksheet.semesterTables({key:value})
    except Exception as e:
        return e


def rangeresult(fromRoll,toRoll):
    try:
        overallresult = generate_marksheet()
        strt = int(fromRoll[6:])
        end  = int(toRoll[6:])
        begn = fromRoll[:6]
        notfound = []
       
        for i in range(strt,end+1):
            rollback = ""
            if i < 9:
                rollback ="0"+str(i)
            else:
                rollback = str(i)        
            roll = begn+rollback
           
            if not roll in overallresult:
                notfound.append(roll)
                continue
            value = overallresult[roll]
            typeofProgram = str(roll[2]+roll[3])
            if typeofProgram=="01":
                if len(value)>9:
                    continue
                btechMarksheet.semesterTables({roll:value})
            else:
                if len(value)>7:
                    continue
                mtechMarksheet.semesterTables({roll:value})
        return notfound
    except Exception as e:
        return e    


